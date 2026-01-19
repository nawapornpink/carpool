from __future__ import annotations

from io import BytesIO
from decimal import Decimal, InvalidOperation
from typing import Iterable, Optional, Any

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.cell.cell import MergedCell


TH_FONT = "TH Sarabun New"

CELL_ODO_L = "A3"
CELL_ODO_S = "B3"
CELL_ODO_E = "G3"
CELL_ODO_EV = "H3"

TABLE_HEADER_TOP = 5
DEFAULT_MAX_ROWS = 28
COLS = "ABCDEFGH"


# =========================
# Helpers
# =========================
def _to_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, Decimal):
        return float(x)
    try:
        s = str(x).strip().replace(",", "")
        if s == "":
            return None
        return float(s)
    except (ValueError, TypeError, InvalidOperation):
        return None


def _set_font(cell, *, bold: bool = False, size: int = 12):
    cell.font = Font(name=TH_FONT, bold=bold, size=size)


def safe_write(ws, addr: str, value: Any):
    """
    ✅ กัน 'MergedCell value is read-only'
    ถ้า addr อยู่ใน merged range แล้วไม่ใช่ช่องซ้ายบน -> เขียนไปที่ช่องซ้ายบนแทน
    """
    cell = ws[addr]
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if addr in rng:
                tl = ws.cell(row=rng.min_row, column=rng.min_col)
                tl.value = value
                return tl
        return cell
    cell.value = value
    return cell


def safe_unmerge_all(ws):
    # unmerge ทั้งหมดแบบไม่จำกัดช่วง กัน template แปลกๆ
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))


def _apply_outline(ws, start_row, end_row, start_col, end_col, thin: Side, thick: Side):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left=thick if c == start_col else thin,
                right=thick if c == end_col else thin,
                top=thick if r == start_row else thin,
                bottom=thick if r == end_row else thin,
            )


def _apply_a4_one_page(ws, *, last_row: int):
    """
    ✅ ล็อกการพิมพ์ให้ "พอดี A4 หน้าเดียว"
    - Paper: A4
    - Portrait
    - Fit to 1 page (Width=1, Height=1)
    - Margin แคบลงนิด เพื่อให้ข้อมูลไม่ล้น
    - Center แนวนอน (ช่วยให้บาลานซ์)
    - Print area: A1:H{last_row}
    """
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    # สำคัญ: ให้ FitToPage ทำงานจริง
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.scale = None  # อย่าล็อก 100% เพราะจะทำให้ FitToPage ไม่คุมเต็มที่

    # บาง Excel ต้อง set ตรงนี้ด้วย
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines = False
    ws.print_options.horizontalCentered = True

    ws.page_margins = PageMargins(
        left=0.30, right=0.30, top=0.40, bottom=0.40, header=0.20, footer=0.20
    )

    ws.print_area = f"A1:H{last_row}"


# =========================
# Main builder
# =========================
def build_fuel_excel(
    template_path: str,
    title_text: str,
    meta_text: str,
    odo_start: int | None = None,
    refills: Optional[Iterable[Any]] = None,
    *,
    max_rows: int = DEFAULT_MAX_ROWS,
    summary_row: int | None = None,
    user_role_label: str = "กดส.1",
    driver_name: str = "",
    controller_name: str = "",
    controller_position: str = "พนัก.6 กดส.1",
    note_line1: str = "ใช้เป็นรถประจำ  กดส.1  และใช้ในราชการเท่านั้น",
    note_line2: str = "(ไม่ได้ใช้เป็นรถประจำตำแหน่ง)",
):
    if refills is None:
        refills = []

    wb = load_workbook(template_path)
    ws = wb.active

    # =========================================================
    # 0) กัน mergedcell จาก template (ชัวร์สุด: unmerge ทั้งหมดก่อน)
    # =========================================================
    safe_unmerge_all(ws)

    # =========================================================
    # 1) Styles
    # =========================================================
    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")
    dotted = Side(style="dotted", color="000000")

    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EFEFEF")

    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    def dotted_underline(cell):
        cell.border = Border(bottom=dotted)

    # =========================================================
    # 2) Header บนสุด
    # =========================================================
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")

    c = safe_write(ws, "A1", title_text)
    _set_font(c, bold=True, size=16)
    c.alignment = align_center

    c = safe_write(ws, "A2", meta_text)
    _set_font(c, bold=False, size=12)
    c.alignment = align_center

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    # =========================================================
    # 3) เลขไมล์ต้นเดือน/สิ้นเดือน
    # =========================================================
    c = safe_write(ws, CELL_ODO_L, "เลขไมล์ต้นเดือน")
    _set_font(c, bold=True, size=12)
    c.alignment = align_left

    if odo_start is not None:
        safe_write(ws, CELL_ODO_S, int(odo_start))
    ws[CELL_ODO_S].number_format = "#,##0"
    ws[CELL_ODO_S].alignment = align_center
    ws[CELL_ODO_S].border = border_thin
    _set_font(ws[CELL_ODO_S], bold=False, size=12)

    c = safe_write(ws, CELL_ODO_E, "เลขไมล์สิ้นเดือน")
    _set_font(c, bold=True, size=12)
    c.alignment = align_right

    ws[CELL_ODO_EV].number_format = "#,##0"
    ws[CELL_ODO_EV].alignment = align_center
    ws[CELL_ODO_EV].border = border_thin
    _set_font(ws[CELL_ODO_EV], bold=False, size=12)

    ws.row_dimensions[3].height = 18

    # =========================================================
    # 4) ✅ หัวตาราง 3 ชั้น (เหมือนรูป)
    # =========================================================
    hdr1 = TABLE_HEADER_TOP
    hdr2 = hdr1 + 1
    hdr3 = hdr1 + 2
    data_start_row = hdr3 + 1

    # เคลียร์พื้นที่หัวตาราง
    for r in range(hdr1, hdr3 + 1):
        ws.row_dimensions[r].height = 18
        for col in COLS:
            cell = ws[f"{col}{r}"]
            cell.value = None
            cell.border = border_thin
            cell.fill = header_fill
            _set_font(cell, bold=True, size=12)
            cell.alignment = align_center_wrap

    # Merge โครงหัวตาราง
    ws.merge_cells(f"A{hdr1}:A{hdr3}")
    ws.merge_cells(f"B{hdr1}:B{hdr3}")
    ws.merge_cells(f"G{hdr1}:G{hdr3}")
    ws.merge_cells(f"H{hdr1}:H{hdr3}")

    ws.merge_cells(f"C{hdr1}:D{hdr1}")
    ws.merge_cells(f"E{hdr1}:F{hdr1}")

    ws.merge_cells(f"A{hdr3}:F{hdr3}")  # เลขไมล์ต้นเดือน

    # ใส่ข้อความ (safe_write ทุกจุด)
    safe_write(ws, f"A{hdr1}", "วันที่\nเติมน้ำมัน")
    safe_write(ws, f"B{hdr1}", "เลขที่ใบส่งจ่าย\nยพ.1")
    safe_write(ws, f"C{hdr1}", "น้ำมันเชื้อเพลิงที่ใช้")
    safe_write(ws, f"E{hdr1}", "น้ำมันหล่อลื่น")
    safe_write(ws, f"G{hdr1}", "เลข กม.\nที่เติม")
    safe_write(ws, f"H{hdr1}", "หมายเหตุ")

    safe_write(ws, f"C{hdr2}", "จำนวนลิตร")
    safe_write(ws, f"D{hdr2}", "จำนวนเงิน")
    safe_write(ws, f"E{hdr2}", "จำนวนลิตร")
    safe_write(ws, f"F{hdr2}", "จำนวนเงิน")

    safe_write(ws, f"A{hdr3}", "เลขไมล์ต้นเดือน")

    ws.row_dimensions[hdr1].height = 26
    ws.row_dimensions[hdr2].height = 20
    ws.row_dimensions[hdr3].height = 18

    # เส้นล่างหนาใต้ hdr2
    for ccol in range(1, 9):
        cell = ws.cell(row=hdr2, column=ccol)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thick)

    # กรอบนอกหัวตารางหนา
    _apply_outline(ws, hdr1, hdr3, 1, 8, thin, thick)

    # =========================================================
    # 5) Table body (ล็อกจำนวนแถว)
    # =========================================================
    table_last_row = data_start_row + max_rows - 1
    for r in range(data_start_row, table_last_row + 1):
        ws.row_dimensions[r].height = 17
        for col in COLS:
            c = ws[f"{col}{r}"]
            c.value = None
            c.border = border_thin
            _set_font(c, bold=False, size=12)
            c.alignment = Alignment(vertical="center", wrap_text=True)

    refills_list = list(refills)[:max_rows]

    last_odo: Optional[int] = None
    for i, f in enumerate(refills_list):
        r = data_start_row + i

        refill_date = getattr(f, "refill_date", None)
        bill_or_yp = getattr(f, "yp_number", "") or ""

        fuel_liters = _to_float(getattr(f, "liters", None))
        fuel_amount = _to_float(getattr(f, "total_price", None))

        lube_liters = _to_float(getattr(f, "lube_liters", None))
        lube_amount = _to_float(getattr(f, "lube_total_price", None))

        odometer = getattr(f, "odometer", None)
        remark = getattr(f, "remark", "") or ""

        safe_write(ws, f"A{r}", refill_date)
        safe_write(ws, f"B{r}", bill_or_yp)
        safe_write(ws, f"C{r}", fuel_liters if fuel_liters is not None else None)
        safe_write(ws, f"D{r}", fuel_amount if fuel_amount is not None else None)

        safe_write(ws, f"E{r}", lube_liters if lube_liters is not None else "-")
        safe_write(ws, f"F{r}", lube_amount if lube_amount is not None else "-")

        safe_write(ws, f"G{r}", odometer)
        safe_write(ws, f"H{r}", remark)

        ws[f"A{r}"].number_format = "dd/mm/yyyy"
        ws[f"C{r}"].number_format = "#,##0.000"
        ws[f"D{r}"].number_format = "#,##0.00"
        ws[f"E{r}"].number_format = "#,##0.000"
        ws[f"F{r}"].number_format = "#,##0.00"
        ws[f"G{r}"].number_format = "#,##0"

        ws[f"A{r}"].alignment = align_center
        ws[f"B{r}"].alignment = align_center
        ws[f"C{r}"].alignment = align_right
        ws[f"D{r}"].alignment = align_right
        ws[f"E{r}"].alignment = align_center if ws[f"E{r}"].value == "-" else align_right
        ws[f"F{r}"].alignment = align_center if ws[f"F{r}"].value == "-" else align_right
        ws[f"G{r}"].alignment = align_right
        ws[f"H{r}"].alignment = align_left

        if odometer not in (None, ""):
            try:
                last_odo = int(odometer)
            except Exception:
                pass

    if last_odo is not None:
        safe_write(ws, CELL_ODO_EV, last_odo)

    # =========================================================
    # 6) แถวรวมใต้ตาราง
    # =========================================================
    total_row = table_last_row + 1
    ws.row_dimensions[total_row].height = 18
    for col in COLS:
        c = ws[f"{col}{total_row}"]
        c.border = border_thin
        _set_font(c, bold=False, size=12)
        c.alignment = align_center

    # =========================================================
    # 7) คำนวณ “ตัวเลขจริง”
    # =========================================================
    fuel_liters_sum = 0.0
    fuel_money_sum = 0.0
    lube_liters_sum = 0.0
    lube_money_sum = 0.0

    for f in refills_list:
        fuel_liters_sum += _to_float(getattr(f, "liters", 0)) or 0.0
        fuel_money_sum += _to_float(getattr(f, "total_price", 0)) or 0.0

        ll = _to_float(getattr(f, "lube_liters", None))
        lm = _to_float(getattr(f, "lube_total_price", None))
        if ll is not None:
            lube_liters_sum += ll
        if lm is not None:
            lube_money_sum += lm

    odo_end = _to_float(ws[CELL_ODO_EV].value)
    odo_start_val = _to_float(ws[CELL_ODO_S].value)

    distance = None
    if odo_end is not None and odo_start_val is not None:
        distance = odo_end - odo_start_val

    avg_km_per_liter = None
    if distance is not None and fuel_liters_sum > 0:
        avg_km_per_liter = distance / fuel_liters_sum

    safe_write(ws, f"C{total_row}", fuel_liters_sum if fuel_liters_sum > 0 else None)
    safe_write(ws, f"D{total_row}", fuel_money_sum if fuel_money_sum > 0 else None)
    safe_write(ws, f"E{total_row}", lube_liters_sum if lube_liters_sum > 0 else None)
    safe_write(ws, f"F{total_row}", lube_money_sum if lube_money_sum > 0 else None)
    safe_write(ws, f"G{total_row}", last_odo if last_odo is not None else None)

    ws[f"C{total_row}"].number_format = "#,##0.000"
    ws[f"D{total_row}"].number_format = "#,##0.00"
    ws[f"E{total_row}"].number_format = "#,##0.000"
    ws[f"F{total_row}"].number_format = "#,##0.00"
    ws[f"G{total_row}"].number_format = "#,##0"

    for addr in [f"C{total_row}", f"D{total_row}", f"E{total_row}", f"F{total_row}", f"G{total_row}"]:
        _set_font(ws[addr], bold=True, size=12)
        ws[addr].alignment = align_right

    # =========================================================
    # 8) สรุปด้านล่าง (เส้นจุด)
    # =========================================================
    r1 = summary_row if isinstance(summary_row, int) else (total_row + 2)
    r2 = r1 + 1
    r3 = r1 + 2
    r4 = r1 + 3

    for rr in (r1, r2, r3, r4):
        ws.row_dimensions[rr].height = 18

    def dash_or_value(v: float):
        return "-" if v <= 0 else v

    safe_write(ws, f"A{r1}", "น้ำมันเชื้อเพลิงรวม")
    _set_font(ws[f"A{r1}"], bold=True)
    safe_write(ws, f"C{r1}", dash_or_value(fuel_liters_sum))
    ws[f"C{r1}"].number_format = "#,##0.000"
    ws[f"C{r1}"].alignment = align_right
    _set_font(ws[f"C{r1}"], bold=True)
    dotted_underline(ws[f"C{r1}"])
    safe_write(ws, f"D{r1}", "ลิตร")
    safe_write(ws, f"E{r1}", "เป็นเงิน")
    safe_write(ws, f"F{r1}", dash_or_value(fuel_money_sum))
    ws[f"F{r1}"].number_format = "#,##0.00"
    ws[f"F{r1}"].alignment = align_right
    _set_font(ws[f"F{r1}"], bold=True)
    dotted_underline(ws[f"F{r1}"])
    safe_write(ws, f"G{r1}", "บาท")

    safe_write(ws, f"A{r2}", "น้ำมันหล่อลื่นรวม")
    _set_font(ws[f"A{r2}"], bold=True)
    safe_write(ws, f"C{r2}", dash_or_value(lube_liters_sum))
    ws[f"C{r2}"].number_format = "#,##0.000"
    ws[f"C{r2}"].alignment = align_right
    _set_font(ws[f"C{r2}"], bold=True)
    dotted_underline(ws[f"C{r2}"])
    safe_write(ws, f"D{r2}", "ลิตร")
    safe_write(ws, f"E{r2}", "เป็นเงิน")
    safe_write(ws, f"F{r2}", dash_or_value(lube_money_sum))
    ws[f"F{r2}"].number_format = "#,##0.00"
    ws[f"F{r2}"].alignment = align_right
    _set_font(ws[f"F{r2}"], bold=True)
    dotted_underline(ws[f"F{r2}"])
    safe_write(ws, f"G{r2}", "บาท")

    safe_write(ws, f"A{r3}", "ระยะทางการใช้รถยนต์ในรอบเดือน")
    _set_font(ws[f"A{r3}"], bold=True)
    safe_write(ws, f"C{r3}", distance if distance is not None else "")
    ws[f"C{r3}"].number_format = "#,##0.00"
    ws[f"C{r3}"].alignment = align_right
    _set_font(ws[f"C{r3}"], bold=True)
    dotted_underline(ws[f"C{r3}"])
    safe_write(ws, f"D{r3}", "กม.")

    safe_write(ws, f"A{r4}", "เฉลี่ยการใช้น้ำมันเชื้อเพลิง")
    _set_font(ws[f"A{r4}"], bold=True)
    safe_write(ws, f"C{r4}", avg_km_per_liter if avg_km_per_liter is not None else "")
    ws[f"C{r4}"].number_format = "#,##0.00"
    ws[f"C{r4}"].alignment = align_right
    _set_font(ws[f"C{r4}"], bold=True)
    dotted_underline(ws[f"C{r4}"])
    safe_write(ws, f"D{r4}", "กม./ลิตร")

    # =========================================================
    # 9) ลายเซ็น + หมายเหตุ
    # =========================================================
    sig_row = r4 + 2
    ws.row_dimensions[sig_row].height = 18
    ws.row_dimensions[sig_row + 1].height = 18

    safe_write(ws, f"A{sig_row}", f"ลงชื่อ {user_role_label}")
    ws[f"A{sig_row}"].alignment = align_left
    _set_font(ws[f"A{sig_row}"], bold=False)

    ws[f"B{sig_row}"].value = ""
    ws[f"C{sig_row}"].value = ""
    dotted_underline(ws[f"B{sig_row}"])
    dotted_underline(ws[f"C{sig_row}"])
    safe_write(ws, f"D{sig_row}", "(ผู้ใช้รถยนต์)")
    _set_font(ws[f"D{sig_row}"], bold=False)

    safe_write(ws, f"E{sig_row}", "ลงชื่อ")
    _set_font(ws[f"E{sig_row}"], bold=False)

    ws[f"F{sig_row}"].value = ""
    ws[f"G{sig_row}"].value = ""
    dotted_underline(ws[f"F{sig_row}"])
    dotted_underline(ws[f"G{sig_row}"])
    safe_write(ws, f"H{sig_row}", "(ผู้ควบคุมการใช้รถยนต์/พหน.)")
    _set_font(ws[f"H{sig_row}"], bold=False)

    name_row = sig_row + 1
    ws.merge_cells(f"A{name_row}:D{name_row}")
    ws.merge_cells(f"E{name_row}:H{name_row}")
    safe_write(ws, f"A{name_row}", f"( {driver_name} )" if driver_name else "(                         )")
    safe_write(
        ws,
        f"E{name_row}",
        f"( {controller_name} )   {controller_position}" if controller_name else f"(                         )   {controller_position}",
    )
    ws[f"A{name_row}"].alignment = align_center
    ws[f"E{name_row}"].alignment = align_center
    _set_font(ws[f"A{name_row}"], bold=False)
    _set_font(ws[f"E{name_row}"], bold=False)

    note_row = name_row + 2
    ws.row_dimensions[note_row].height = 18
    ws.row_dimensions[note_row + 1].height = 18

    safe_write(ws, f"A{note_row}", "หมายเหตุ:-")
    _set_font(ws[f"A{note_row}"], bold=True)
    ws[f"A{note_row}"].alignment = align_left

    ws.merge_cells(f"B{note_row}:H{note_row}")
    safe_write(ws, f"B{note_row}", note_line1)
    ws[f"B{note_row}"].alignment = align_left
    _set_font(ws[f"B{note_row}"], bold=False)

    ws.merge_cells(f"B{note_row + 1}:H{note_row + 1}")
    safe_write(ws, f"B{note_row + 1}", note_line2)
    ws[f"B{note_row + 1}"].alignment = align_left
    _set_font(ws[f"B{note_row + 1}"], bold=False)

    # =========================================================
    # 10) กรอบตารางรวม
    # =========================================================
    _apply_outline(ws, hdr1, total_row, 1, 8, thin, thick)

    # =========================================================
    # 11) ความกว้างคอลัมน์ (ปรับให้ A4 พอดีขึ้น)
    # =========================================================
    # เดิมค่อนข้างกว้าง ทำให้บางเครื่องพิมพ์ล้นขวาได้
    widths = [10.5, 12, 12.5, 11.5, 12.5, 11.5, 11, 13.5]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # =========================================================
    # 12) Page setup: A4 หน้าเดียว + Print area
    # =========================================================
    last_row_for_print = note_row + 2
    _apply_a4_one_page(ws, last_row=last_row_for_print)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
